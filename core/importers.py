"""
数据底座 — Excel导入器
支持恩同的五种数据格式：配件清单/供应商进价/客户名单/报价留底
"""

import sqlite3
import json
import openpyxl
from pathlib import Path
import sys
sys.path.insert(0, str(Path(__file__).parent.parent))
from database import get_db, rebuild_fts


def import_parts(excel_path: str, sheet_name: str = None) -> dict:
    """
    导入配件清单Excel
    期望列：产品线, OE号, 替代OE号, 中文品名, 俄文品名, 英文品名,
           品牌/渠道, 供货号, 牌价, 适配发动机, 适配车型, 排放标准,
           单位, 定价模式, 库存一口价, 含税进价, 备注
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    
    conn = get_db()
    count = 0
    errors = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1] or str(row[1]).strip() == '':  # 跳过空行（需要OE号）
            continue
        
        try:
            conn.execute("""
                INSERT OR REPLACE INTO parts (
                    oe_number, alt_oe_numbers, name_cn, name_ru, name_en,
                    brand_channel, supply_number, list_price,
                    engine_model, vehicle_model, emission_std,
                    unit, pricing_mode, fixed_stock_price, cost_with_tax,
                    product_line, notes
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                str(row[1]).strip(),           # OE号
                str(row[2] or ''),              # 替代OE号
                str(row[3] or ''),              # 中文品名
                str(row[4] or ''),              # 俄文品名
                str(row[5] or ''),              # 英文品名
                str(row[6] or ''),              # 品牌
                str(row[7] or ''),              # 供货号
                float(row[8] or 0),             # 牌价
                str(row[9] or ''),              # 适配发动机
                str(row[10] or ''),             # 适配车型
                str(row[11] or ''),             # 排放标准
                str(row[12] or 'PC'),           # 单位
                str(row[13] or 'STANDARD'),     # 定价模式
                float(row[14] or 0),            # 库存一口价
                float(row[15] or 0),            # 含税进价
                str(row[0] or ''),              # 产品线
                str(row[16] or ''),             # 备注
            ))
            count += 1
        except Exception as e:
            errors.append(f"行{row[0]}: {str(e)[:80]}")
    
    conn.commit()
    rebuild_fts()  # 重建全文索引
    conn.close()
    
    return {"imported": count, "errors": errors}


def import_suppliers(excel_path: str) -> dict:
    """导入供应商进价表"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    conn = get_db()
    count = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        try:
            conn.execute("""
                INSERT OR REPLACE INTO suppliers (name, brands, contact, payment_terms)
                VALUES (?, ?, ?, ?)
            """, (
                str(row[0]).strip(),
                str(row[1] or ''),
                str(row[2] or ''),
                str(row[3] or ''),
            ))
            count += 1
        except:
            pass
    
    conn.commit()
    conn.close()
    return {"imported": count}


def import_customers(excel_path: str) -> dict:
    """导入客户名单"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    conn = get_db()
    count = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        try:
            conn.execute("""
                INSERT OR REPLACE INTO customers (
                    name_cn, name_en, country, region, star_level,
                    annual_purchase, preferred_trade, preferred_payment,
                    payment_punctuality, tags, notes
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                str(row[0] or ''),
                str(row[1] or ''),
                str(row[2] or ''),
                str(row[3] or ''),
                int(row[4] or 1),
                float(row[5] or 0),
                str(row[6] or 'FOB'),
                str(row[7] or 'prepaid'),
                str(row[8] or ''),
                str(row[9] or '[]'),
                str(row[10] or ''),
            ))
            count += 1
        except:
            pass
    
    conn.commit()
    conn.close()
    return {"imported": count}


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("用法: python excel_importer.py <配件Excel路径>")
        sys.exit(1)
    
    result = import_parts(sys.argv[1])
    print(f"导入完成: {result['imported']} 条配件")
    if result['errors']:
        print(f"错误: {len(result['errors'])} 条")
        for e in result['errors'][:5]:
            print(f"  - {e}")
