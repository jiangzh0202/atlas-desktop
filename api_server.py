"""
擎天·API 服务 — Flask REST API (端口 3092)
给 Node.js 操作台调用的报价引擎接口
"""
import sys, os, json, io
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent))

from flask import Flask, request, jsonify, send_file, g
from flask_cors import CORS

import sys, os
sys.path.insert(0, os.path.dirname(__file__))

from config import cfg, get_plans, get_brand_rules, get_approval_flow, get_price_floors, get_dimension, calc_procurement_units, sku_per_unit
from middleware import require_auth, require_role, sign_jwt, verify_jwt
from puzzler_engine import (find_leads, background_check, generate_email, full_pipeline, match_analysis, configure_smtp, send_email_via_smtp, send_bulk_emails, SMTP_CONFIG, PRODUCT_LINES, INFO_SOURCES)
from pipeline import get_pipeline  # 硬管道骨架
from review_criteria import ReviewCriteria  # 审查标准保护层
from email_utils import check_spam_score, build_html_email
from models import (
    init_db, ensure_admin, verify_login, get_user, list_users, create_user, delete_user,
    role_can, get_role_label, log_usage, get_usage, ROLES, DB_PATH
)

# Init DB
ensure_admin()
app = Flask(__name__)
CORS(app)

# ─── 健康检查 ───
@app.route("/api/health")
def health():
    return jsonify({"ok": True, "service": "atlas-api", "version": "2.0"})

@app.route("/api/login", methods=["POST"])
def api_login():
    data = request.get_json() or {}
    email = data.get("email", "").strip()
    password = data.get("password", "")
    if not email or not password:
        return jsonify({"ok": False, "error": "请输入邮箱和密码"}), 400
    user = verify_login(email, password)
    if not user:
        return jsonify({"ok": False, "error": "邮箱或密码错误"}), 401
    from middleware import sign_jwt
    import time
    token = sign_jwt({
        "sub": user["id"], "user_id": user["id"],
        "role": user.get("role", "user"),
        "company_id": user.get("company_id", ""),
        "iat": int(time.time()), "exp": int(time.time()) + 86400 * 7
    })
    return jsonify({"ok": True, "token": token, "user": {
        "id": user["id"], "name": user.get("name",""), "email": email,
        "role": user.get("role","user"), "company_id": user.get("company_id","")
    }})

@app.route("/api/register", methods=["POST"])
def api_register():
    data = request.get_json() or {}
    email = data.get("email", "").strip()
    password = data.get("password", "")
    name = data.get("name", email.split("@")[0] if "@" in email else "")
    company = data.get("company", "")
    if not email or not password:
        return jsonify({"ok": False, "error": "请输入邮箱和密码"}), 400
    if "@" not in email or len(password) < 4:
        return jsonify({"ok": False, "error": "邮箱格式错误或密码过短（至少4位）"}), 400
    # Check existing
    existing = verify_login(email, password)
    # Actually check if just email exists (wrong password also means exists)
    try:
        cur = db.cursor()
        cur.execute("SELECT id FROM users WHERE email = ?", (email,))
        if cur.fetchone():
            return jsonify({"ok": False, "error": "该邮箱已注册，请直接登录"}), 409
    except:
        pass
    import uuid
    uid = "U-" + str(uuid.uuid4())[:12]
    try:
        create_user(email=email, password=password, name=name, role="user", company_id=company or "COMPANY-DEFAULT")
        return jsonify({"ok": True, "message": "注册成功"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


# ─── 配件搜索 ───
@app.route("/api/parts")
@require_role("quote")
def search_parts():
    q = request.args.get("q", "")
    if not q:
        return jsonify({"ok": False, "error": "缺少搜索词"})
    try:
        from core import search_parts as sp
        results, total = sp(q, limit=20)
        return jsonify({"ok": True, "data": results, "count": total})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e), "data": []})

# ─── 解析上传的Excel询盘 ───
@app.route("/api/parse", methods=["POST"])
@require_role("quote")
def parse_inquiry():
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "请上传Excel文件"})
    
    file = request.files["file"]
    if file.filename == "":
        return jsonify({"ok": False, "error": "文件名为空"})
    
    try:
        # 保存临时文件
        tmp_path = f"/tmp/atlas_upload_{os.getpid()}.xlsx"
        file.save(tmp_path)
        
        # 用现有的 EnTongWorkbook 解析器
        from atlas.parsers.enong_workbook import EnTongWorkbook
        wb = EnTongWorkbook(tmp_path)
        
        inquiries = wb.parse_inquiry()
        parts = wb.parse_worksheet()
        
        # 组装 items 列表（询盘 + 匹配的配件信息合并）
        items = []
        for inq in inquiries:
            item = {
                "oe_number": inq.oe_number,
                "name_cn": getattr(inq, "name_cn", ""),
                "name_ru": getattr(inq, "name_ru", ""),
                "quantity": inq.quantity,
                "matched": False
            }
            # 尝试匹配报价留底中的配件
            for p in parts:
                if getattr(p, "oe_number", "") == inq.oe_number or getattr(p, "reference_oe", "") == inq.oe_number:
                    item["list_price"] = getattr(p, "list_price", 0)
                    item["brand_channel"] = getattr(p, "brand_channel", "")
                    item["name_cn"] = getattr(p, "name_cn", "") or item["name_cn"]
                    item["matched"] = True
                    break
            items.append(item)
        
        result = {"items": items, "count": len(items)}
        
        os.remove(tmp_path)
        return jsonify({"ok": True, "data": result})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)})

# ─── 报价计算 ───
@app.route("/api/quote", methods=["POST"])
@require_role("quote")
def calculate_quote():
    try:
        data = request.get_json(force=True)
        items = data.get("items", [])
        customer_id = data.get("customer_id", "")
        trade_term = data.get("trade_term", "FOB")
        payment_term = data.get("payment_term", "prepaid")
        
        if not items:
            return jsonify({"ok": False, "error": "询盘数据为空"})
        
        import asyncio
        from agents.quotation_agent import QuotationAgent
        
        agent = QuotationAgent()
        loop = asyncio.new_event_loop()
        result = loop.run_until_complete(
            agent.process(items, customer_id, trade_term, payment_term)
        )
        loop.close()
        
        # 获取审计追踪
        from ledger.trace import audit
        trace = audit.get_trace(result["quotation_id"])
        
        return jsonify({
            "ok": result.get("ok", True),
            "data": result,
            "trace": trace,
            "revised_count": result.get("revised_count", 0),
            "audit_trail_length": result.get("audit_trail_length", len(trace)),
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)})


# ─── 报价 Excel 导出 ───
@app.route("/api/quote/<quote_id>/export")
@require_role("quote")
def export_quote_excel(quote_id):
    """
    GET /api/quote/<quote_id>/export — 导出报价单为 Excel
    返回 .xlsx 文件，可直接发客户
    """
    try:
        from ledger.trace import audit
        trace = audit.get_trace(quote_id)
        if not trace:
            return jsonify({"ok": False, "error": f"报价 {quote_id} 不存在"})

        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
        from openpyxl.utils import get_column_letter
        import io

        wb = Workbook()
        ws = wb.active
        ws.title = "报价单"

        # ─── 样式 ───
        header_font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="7ECF5E", end_color="7ECF5E", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        body_font = Font(name="Microsoft YaHei", size=10)
        body_align = Alignment(horizontal="center", vertical="center")
        money_fmt = '#,##0.00'
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        total_font = Font(name="Microsoft YaHei", bold=True, size=12)

        # ─── 标题行 ───
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = f"报价单 — {quote_id}"
        title_cell.font = Font(name="Microsoft YaHei", bold=True, size=14, color="1A1A2E")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # ─── 表头 ───
        headers = ["序号", "OE号", "品名", "品牌", "数量", "单价", "小计"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # ─── 从审计日志重建 items（match.name + price.price/total）───
        match_map = {}   # oe → {name}
        price_items = [] # [{oe, qty, price, total}]
        for entry in trace:
            if entry.get("step") == "match":
                inp = entry.get("input", {})
                oe = inp.get("oe", "").strip()
                if oe:
                    match_map[oe] = inp.get("name", "")
            elif entry.get("step") == "price":
                inp = entry.get("input", {})
                out = entry.get("output", {})
                oe = inp.get("oe", "").strip()
                if oe and out.get("price", 0):
                    price_items.append({
                        "oe": oe,
                        "qty": inp.get("qty", 1),
                        "price": out.get("price", 0),
                        "total": out.get("total", 0),
                    })

        if not price_items:
            return jsonify({
                "ok": False,
                "error": "该报价无明细数据，无法导出。建议重新计算报价后导出。"
            })

        # ─── 数据行 ───
        total_amount = 0
        row = 3
        for i, item in enumerate(price_items, 1):
            oe = item.get("oe", "")
            name = match_map.get(oe, "")
            qty = item.get("qty", 1)
            price = item.get("price", 0)
            subtotal = item.get("total", 0)
            brand = ""  # 审计日志暂不记录 brand

            ws.cell(row=row, column=1, value=i).font = body_font
            ws.cell(row=row, column=1).alignment = body_align
            ws.cell(row=row, column=1).border = thin_border

            ws.cell(row=row, column=2, value=oe).font = body_font
            ws.cell(row=row, column=2).alignment = body_align
            ws.cell(row=row, column=2).border = thin_border

            ws.cell(row=row, column=3, value=name).font = body_font
            ws.cell(row=row, column=3).alignment = Alignment(horizontal="left", vertical="center")
            ws.cell(row=row, column=3).border = thin_border

            ws.cell(row=row, column=4, value=brand).font = body_font
            ws.cell(row=row, column=4).alignment = body_align
            ws.cell(row=row, column=4).border = thin_border

            ws.cell(row=row, column=5, value=int(qty)).font = body_font
            ws.cell(row=row, column=5).alignment = body_align
            ws.cell(row=row, column=5).border = thin_border

            price_cell = ws.cell(row=row, column=6, value=float(price))
            price_cell.font = body_font
            price_cell.alignment = body_align
            price_cell.number_format = money_fmt
            price_cell.border = thin_border

            subtotal_cell = ws.cell(row=row, column=7, value=float(subtotal))
            subtotal_cell.font = body_font
            subtotal_cell.alignment = body_align
            subtotal_cell.number_format = money_fmt
            subtotal_cell.border = thin_border

            total_amount += float(subtotal)
            ws.row_dimensions[row].height = 22
            row += 1

        total_amount = round(total_amount, 2)  # 修正浮点累积误差

        # ─── 合计行 ───
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        total_label = ws.cell(row=row, column=1, value="合计")
        total_label.font = total_font
        total_label.alignment = Alignment(horizontal="right", vertical="center")
        total_label.border = thin_border
        for c in range(2, 6):
            ws.cell(row=row, column=c).border = thin_border

        total_cell = ws.cell(row=row, column=6, value=total_amount)
        total_cell.font = total_font
        total_cell.alignment = body_align
        total_cell.number_format = money_fmt
        total_cell.border = thin_border

        ws.cell(row=row, column=7).border = thin_border
        ws.row_dimensions[row].height = 26

        # ─── 列宽 ───
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 32
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 14

        # ─── 输出 ───
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"报价单_{quote_id}.xlsx"
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)})


# ─── 报价对比 ───
@app.route("/api/quote/compare", methods=["POST"])
@require_role("quote")
def compare_quote():
    """
    POST /api/quote/compare — 报价对比
    请求: {
        "quote_result": {...},           # process() 返回的完整报价结果
        "original_excel_path": "/path/to/报价留底.xlsx"  # 原始 Excel 路径
    }
    返回: {"ok": true, "data": {matched, diff_lines, accuracy_pct, ...}}
    """
    try:
        data = request.get_json(force=True)
        quote_result = data.get("quote_result")
        original_excel_path = data.get("original_excel_path", "")
        
        if not quote_result:
            return jsonify({"ok": False, "error": "缺少 quote_result"})
        if not original_excel_path:
            return jsonify({"ok": False, "error": "缺少 original_excel_path"})
        
        from agents.quotation_agent import QuotationAgent
        result = QuotationAgent.compare_with_original(quote_result, original_excel_path)
        return jsonify(result)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)})

# ─── 客户列表 ───
@app.route("/api/customers")
@require_role("quote")
def list_customers():
    try:
        from core import get_db
        db = get_db()
        rows = db.execute("SELECT id, name_cn, name_en, country, region, star_level, annual_purchase, preferred_trade, preferred_payment FROM customers LIMIT 100").fetchall()
        customers = [dict(r) for r in rows]
        return jsonify({"ok": True, "data": customers, "count": len(customers)})
    except Exception as e:
        return jsonify({"ok": True, "data": [], "count": 0, "note": f"客户表为空: {e}"})

# ─── 规则面板 ───
@app.route("/api/rules", methods=["GET", "POST"])
@require_role("manage")
def handle_rules():
    import os, shutil, time
    agent = request.args.get("agent", "quoter").strip()
    AGENT_RULES = {
        "quoter": "rules.quoter.md",
        "developer": "rules.developer.md",
        "buyer": "rules.buyer.md",
        "image": "rules.image.md",
        "stock": "rules.stock.md",
        "customs": "rules.customs.md",
    }
    rules_file = AGENT_RULES.get(agent, "rules.md")
    rules_path = os.path.join(os.path.dirname(__file__), "data", rules_file)
    backup_dir = os.path.join(os.path.dirname(__file__), "data", "rules_backups")
    os.makedirs(backup_dir, exist_ok=True)
    
    if request.method == "GET":
        raw = open(rules_path, encoding="utf-8").read() if os.path.exists(rules_path) else "# 报价规则"
        return jsonify({"ok": True, "raw": raw, "size": len(raw)})
    
    data = request.get_json(force=True)
    content = data.get("content", "")
    
    if os.path.exists(rules_path):
        ts = time.strftime("%Y%m%d_%H%M%S")
        shutil.copy2(rules_path, os.path.join(backup_dir, f"rules_{ts}.md"))
    
    with open(rules_path, "w", encoding="utf-8") as f:
        f.write(content)
    
    backups = sorted([f for f in os.listdir(backup_dir) if f.endswith(".md")], reverse=True)[:10]
    return jsonify({"ok": True, "message": "saved", "size": len(content), "backups": backups})

@app.route("/api/rules/versions")
@require_role("manage")
def list_rule_versions():
    import time
    import os
    d = os.path.join(os.path.dirname(__file__), "data", "rules_backups")
    os.makedirs(d, exist_ok=True)
    vers = []
    for f in sorted(os.listdir(d), reverse=True):
        if f.endswith(".md"):
            p = os.path.join(d, f)
            vers.append({"filename": f, "size": os.path.getsize(p), "modified": time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(os.path.getmtime(p)))})
    return jsonify({"ok": True, "data": vers[:20], "count": len(vers)})

@app.route("/api/rules/restore/<filename>")
@require_role("manage")
def restore_rule_version(filename):
    import time
    import os, shutil
    d = os.path.join(os.path.dirname(__file__), "data", "rules_backups")
    rp = os.path.join(os.path.dirname(__file__), "data", "rules.md")
    bp = os.path.join(d, filename)
    if not os.path.exists(bp):
        return jsonify({"ok": False, "error": "version not found"})
    if os.path.exists(rp):
        shutil.copy2(rp, os.path.join(d, f"rules_before_restore_{time.strftime('%Y%m%d_%H%M%S')}.md"))
    shutil.copy2(bp, rp)
    raw = open(rp, encoding="utf-8").read()
    return jsonify({"ok": True, "raw": raw, "message": f"Restored to {filename}"})

# ─── 训练 — 上传历史报价Excel → AI审视提取规则 → 更新记忆 ───
# ─── 知识库上传 — 上传参考文件到指定Agent的知识库 ───
@app.route("/api/kb/upload", methods=["POST"])
@require_role("quote")
def kb_upload():
    """
    上传参考文件到指定Agent的知识库目录
    Body: multipart/form-data with 'file' and 'agent' fields
    Returns: {ok, filename, path}
    """
    try:
        if 'file' not in request.files:
            return jsonify({"ok": False, "error": "请上传文件"})
        f = request.files['file']
        agent = request.form.get('agent', 'quoter').strip()
        # Security: only allow alphanumeric agent names
        import re
        if not re.match(r'^[a-z]+$', agent):
            return jsonify({"ok": False, "error": "无效的agent名称"})
        
        kb_dir = os.path.join(os.path.dirname(__file__), "data", "kb", agent)
        os.makedirs(kb_dir, exist_ok=True)
        
        # Safe filename
        safe_name = re.sub(r'[^a-zA-Z0-9_.\-]', '_', f.filename)
        filepath = os.path.join(kb_dir, safe_name)
        f.save(filepath)
        
        return jsonify({"ok": True, "filename": safe_name, "path": filepath, "agent": agent})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


# ─── 知识库列表 — 返回已上传的知识库文件清单 ───
@app.route("/api/kb/list")
@require_role("quote")
def kb_list():
    """返回所有Agent的知识库文件列表"""
    try:
        agent_filter = request.args.get("agent", "").strip()
        kb_root = os.path.join(os.path.dirname(__file__), "data", "kb")
        files = []
        if os.path.exists(kb_root):
            for agent_dir in sorted(os.listdir(kb_root)):
                if agent_filter and agent_dir != agent_filter:
                    continue
                agent_path = os.path.join(kb_root, agent_dir)
                if os.path.isdir(agent_path):
                    for fname in sorted(os.listdir(agent_path)):
                        fpath = os.path.join(agent_path, fname)
                        if os.path.isfile(fpath):
                            st = os.stat(fpath)
                            files.append({
                                "name": fname,
                                "agent": agent_dir,
                                "size": st.st_size,
                                "modified": st.st_mtime,
                                "path": f"data/kb/{agent_dir}/{fname}"
                            })
        return jsonify({"ok": True, "files": files, "total": len(files)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route("/api/train", methods=["POST"])
@require_role("quote")
def train_from_history():
    """
    客户上传历史报价Excel → AI按审视表逐维度提取定价规则 → 写入 rules.{agent}.md（记忆）
    ?agent=quoter|developer|image|stock|customs — 指定训练哪个数字员工
    Body: multipart/form-data with 'file' (xlsx) and optional 'customer_name'
    Returns: training report with extracted patterns
    """
    agent = request.args.get("agent", "quoter").strip()
    # Map agent to rules file
    AGENT_RULES = {
        "quoter": "rules.quoter.md",
        "developer": "rules.developer.md",
        "buyer": "rules.buyer.md",
        "image": "rules.image.md",
        "stock": "rules.stock.md",
        "customs": "rules.customs.md",
    }
    rules_filename = AGENT_RULES.get(agent, "rules.md")
    try:
        import tempfile, re
        from openpyxl import load_workbook
        
        if 'file' not in request.files:
            return jsonify({"ok": False, "error": "请上传 xlsx 文件"})
        
        f = request.files['file']
        if not f.filename.endswith(('.xlsx', '.xls')):
            return jsonify({"ok": False, "error": "仅支持 .xlsx 格式"})
        
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        f.save(tmp.name)
        wb = load_workbook(tmp.name, data_only=True)
        os.unlink(tmp.name)
        
        # ─── Agent-adaptive column detection ───
        AGENT_COLUMNS = {
            "quoter":    ['折扣', '单价', '数量', '牌价', '折后', '折扣率'],
            "developer": ['客户', '国家', '行业', '来源', '联系人', '邮箱', '电话', '采购'],
            "buyer":     ['供应商', '报价', '单价', '品牌', '货期', '付款', '税点'],
            "stock":     ['库存', '安全库存', '周转', '入库', '出库', 'SKU', '品名'],
            "customs":   ['HS', '海关', '品名', '申报', '报关', '税率', '原产'],
            "image":     ['零件', '型号', 'OE号', '识别', '标签'],
        }
        keywords = AGENT_COLUMNS.get(agent, AGENT_COLUMNS['quoter'])
        
        # Find the best-matching sheet
        target_sheet = None
        header_row = None
        col_map = {}
        best_score = 0
        
        for sname in wb.sheetnames:
            ws = wb[sname]
            for r in range(1, min(10, ws.max_row + 1)):
                header = []
                for c in range(1, ws.max_column + 1):
                    val = str(ws.cell(r, c).value or '').strip()
                    header.append(val)
                header_str = ' '.join(header)
                score = sum(1 for k in keywords if k in header_str)
                if score > best_score:
                    best_score = score
                    target_sheet = sname
                    header_row = r
                    col_map = {}
                    for c in range(1, ws.max_column + 1):
                        val = str(ws.cell(r, c).value or '').strip()
                        if val:
                            col_map[val] = c
            if target_sheet and best_score >= len(keywords) // 2 + 1:
                break
        
        if not target_sheet or best_score == 0:
            # Fallback: use first sheet, treat all columns as data
            target_sheet = wb.sheetnames[0]
            ws = wb[target_sheet]
            header_row = 1
            col_map = {}
            for c in range(1, ws.max_column + 1):
                val = str(ws.cell(1, c).value or '').strip() or f'Column_{c}'
                col_map[val] = c
        
        # Extract all data rows
        records = []
        ws = wb[target_sheet]
        col_count = len(col_map)
        for r in range(header_row + 1, ws.max_row + 1):
            row = {}
            has_data = False
            for label, col in col_map.items():
                val = ws.cell(r, col).value
                row[label] = val
                if val is not None:
                    has_data = True
            if has_data:
                records.append({'row': r, **row})
        
        # ─── Agent-specific pattern analysis ───
        patterns = {}
        pattern_type = "模式"
        
        if agent == "quoter":
            # Extract discount patterns
            discount_records = []
            for rec in records:
                discount = None
                qty = None
                for key, val in rec.items():
                    kv = str(key).lower()
                    if val is None:
                        continue
                    try:
                        v = float(str(val).replace(',','').replace(' ',''))
                    except:
                        continue
                    if '数量' in kv or 'кол' in kv or 'qty' in kv:
                        qty = int(v)
                    elif ('折扣' in kv or '率' in kv) and '系数' not in kv:
                        discount = v
                if discount is not None:
                    rounded = round(discount, -1) if discount > 10 else round(discount)
                    key = f"折扣{rounded}%"
                    if key not in patterns:
                        patterns[key] = {'count': 0, 'values': []}
                    patterns[key]['count'] += 1
                    patterns[key]['values'].append(discount)
            pattern_type = "折扣模式"
        
        elif agent == "developer":
            # Extract customer/country/industry distributions
            for col_name in ['国家', '行业', '来源', '客户']:
                cats = {}
                for rec in records:
                    for key, val in rec.items():
                        if col_name in str(key) and val:
                            v = str(val).strip()
                            if v and v != 'None':
                                cats[v] = cats.get(v, 0) + 1
                if cats:
                    top = sorted(cats.items(), key=lambda x: x[1], reverse=True)[:10]
                    for name, cnt in top:
                        patterns[f"{col_name}:{name}"] = {'count': cnt, 'values': []}
            pattern_type = "客户分布"
        
        elif agent == "stock":
            # Extract stock quantity ranges
            for rec in records:
                for key, val in rec.items():
                    kv = str(key).lower()
                    if val is None:
                        continue
                    try:
                        v = int(float(str(val).replace(',','').replace(' ','')))
                    except:
                        continue
                    if any(k in kv for k in ['库存', '数量', '安全']):
                        if v < 10: bucket = "0-9"
                        elif v < 100: bucket = "10-99"
                        elif v < 1000: bucket = "100-999"
                        else: bucket = "1000+"
                        key = f"库存:{bucket}"
                        if key not in patterns:
                            patterns[key] = {'count': 0, 'values': []}
                        patterns[key]['count'] += 1
                        patterns[key]['values'].append(v)
            pattern_type = "库存分布"
        
        elif agent == "buyer":
            # Extract supplier + price patterns
            for rec in records:
                supplier = None
                price = None
                for key, val in rec.items():
                    if val is None:
                        continue
                    kv = str(key).lower()
                    if '供应商' in kv or '供货' in kv:
                        supplier = str(val).strip()
                    elif '报价' in kv or '单价' in kv or '价格' in kv:
                        try:
                            price = float(str(val).replace(',','').replace(' ',''))
                        except:
                            pass
                if supplier and price is not None:
                    key = f"供应商:{supplier}"
                    if key not in patterns:
                        patterns[key] = {'count': 0, 'values': []}
                    patterns[key]['count'] += 1
                    patterns[key]['values'].append(price)
            # Also count by price range
            price_ranges = {}
            for rec in records:
                for key, val in rec.items():
                    if val is None:
                        continue
                    try:
                        p = float(str(val).replace(',','').replace(' ',''))
                    except:
                        continue
                    if any(k in str(key).lower() for k in ['报价', '单价', '价格']):
                        if p < 100: bucket = "0-99"
                        elif p < 1000: bucket = "100-999"
                        elif p < 5000: bucket = "1000-4999"
                        else: bucket = "5000+"
                        k2 = f"价格区间:{bucket}"
                        patterns[k2] = patterns.get(k2, {'count': 0, 'values': []})
                        patterns[k2]['count'] += 1
                        patterns[k2]['values'].append(p)
            pattern_type = "供应商报价分布"
        
        elif agent == "customs":
            # Extract HS code patterns
            for rec in records:
                for key, val in rec.items():
                    kv = str(key).lower()
                    if val is None:
                        continue
                    if any(k in kv for k in ['hs', '海关', '编码', 'code']):
                        code = str(val).strip()[:4]
                        key = f"HS前4位:{code}"
                        if key not in patterns:
                            patterns[key] = {'count': 0, 'values': []}
                        patterns[key]['count'] += 1
            pattern_type = "HS编码分布"
        
        else:
            # Generic: count columns and data distribution
            patterns['数据行数'] = {'count': len(records), 'values': []}
            patterns['列数'] = {'count': col_count, 'values': []}
            patterns['工作表'] = {'count': 1, 'values': [target_sheet] if isinstance(target_sheet, str) else []}
            pattern_type = "数据概览"
        
        # ─── Generate training report ───
        agent_cn = {"developer":"客户开发员","quoter":"报价员","buyer":"采购员","image":"图片识零件员","stock":"库存管理员","customs":"清关助理"}
        agent_name = agent_cn.get(agent, agent)
        
        report_lines = [
            f"## {agent_name} 训练报告 · {f.filename}",
            f"训练时间: {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M')}",
            f"数据源: {target_sheet} (共{len(wb.sheetnames)}个工作表)",
            f"提取数据行数: {len(records)}",
            f"识别列数: {len(col_map)}",
            f"",
            f"### 发现的{pattern_type}",
        ]
        
        for key, data in sorted(patterns.items(), key=lambda x: x[1]['count'], reverse=True):
            vals = data.get('values', [])
            extra = ''
            if vals:
                try:
                    avg = sum(vals) / len(vals)
                    extra = f', 平均值={avg:.1f}'
                except:
                    pass
            report_lines.append(f"- {key}: 出现{data['count']}次{extra}")
        
        report = '\n'.join(report_lines)
        
        # Append to rules.md as training memory
        rules_path = os.path.join(os.path.dirname(__file__), "data", rules_filename)
        if os.path.exists(rules_path):
            with open(rules_path, "a", encoding="utf-8") as rf:
                rf.write(f"\n\n{report}\n")
        
        return jsonify({
            "ok": True,
            "message": f"训练完成。从 {target_sheet} 提取了 {len(records)} 条报价记录",
            "record_count": len(records),
            "patterns_found": len(patterns),
            "report": report,
            "sheet_used": target_sheet,
            "total_sheets": len(wb.sheetnames),
        })
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)})

@app.route("/api/trace/<quote_id>")
@require_role("quote")
def get_trace(quote_id):
    from ledger.trace import audit
    trace = audit.get_trace(quote_id)
    return jsonify({"ok": True, "data": trace, "count": len(trace)})

# ─── 报价历史 ───
@app.route("/api/history")
@require_role("quote")
def list_history():
    """返回完整报价历史（含金额/日期/配件数/状态）"""
    try:
        from core import get_db
        conn = get_db()
        quotes = conn.execute("""
            SELECT q.id, q.created_at, q.total_amount, q.status,
                   q.customer_name, q.trade_term, q.payment_term,
                   COUNT(ql.id) as line_count
            FROM quotations q
            LEFT JOIN quotation_lines ql ON ql.quotation_id = q.id
            GROUP BY q.id
            ORDER BY q.created_at DESC
            LIMIT 50
        """).fetchall()
        conn.close()

        data = []
        for q in quotes:
            data.append({
                "id": q["id"],
                "created_at": q["created_at"] or "",
                "total_amount": q["total_amount"] or 0,
                "item_count": q["line_count"] or 0,
                "status": q["status"] or "draft",
                "customer_name": q["customer_name"] or "",
                "trade_term": q["trade_term"] or "",
                "payment_term": q["payment_term"] or "",
                "line_count": q["line_count"] or 0,
            })

        # 回退：如果 quotations 表为空，用审计日志
        if not data:
            from ledger.trace import audit
            quote_ids = audit.get_all_quotes()
            data = [{"id": qid, "created_at": "", "total_amount": 0,
                     "item_count": 0, "status": "archived",
                     "customer_name": "", "trade_term": "", "payment_term": "",
                     "line_count": 0} for qid in quote_ids]

        return jsonify({"ok": True, "data": data, "count": len(data)})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)})

# ─── 库存查询 ───
@app.route("/api/stock")
@require_role("quote")
def get_stock_api():
    try:
        from core import get_stock
        data, alert_count = get_stock()
        return jsonify({"ok": True, "data": data, "count": len(data), "alerts": alert_count})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

# ─── 库存预警 ───
@app.route("/api/stock/alerts")
@require_role("quote")
def get_stock_alerts_api():
    try:
        from core import get_stock_alerts
        data = get_stock_alerts()
        return jsonify({"ok": True, "data": data, "count": len(data)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

# ─── 事件链触发（模拟报价完成 → 减库存 → 预警 → 补货建议）───
@app.route("/api/trigger-event", methods=["POST"])
@require_role("quote")
def trigger_event():
    """
    模拟报价完成事件，触发完整事件链:
    报价完成 → warehouse减库存 → 低于安全线 → purchasing生成补货建议
    参数: {"items": [{"oe_number": "xxx", "quantity": N}], "quotation_id": "可选"}
    返回: 事件链结果（减库存结果 + 预警列表 + 采购建议）
    """
    try:
        data = request.get_json(force=True)
        items = data.get("items", [])
        quote_id = data.get("quotation_id", "")

        if not items:
            return jsonify({"ok": False, "error": "items 为空"})

        import asyncio
        import time
        import uuid

        if not quote_id:
            quote_id = f"EVENT-{int(time.time())}-{uuid.uuid4().hex[:6]}"

        # 构建事件
        event = {
            "id": uuid.uuid4().hex[:8],
            "event_type": "quotation.completed",
            "payload": {
                "quotation_id": quote_id,
                "items": items,
                "item_count": len(items)
            },
            "timestamp": time.time()
        }

        # 在临时 event loop 中运行事件链
        from bus.eventbus import EventBus, QUOTATION_COMPLETED
        from agents.warehouse_agent import WarehouseAgent
        from agents.purchasing_agent import PurchasingAgent

        warehouse = WarehouseAgent()
        purchasing = PurchasingAgent()

        async def run_chain():
            # 重置单例避免 event loop 绑定问题
            EventBus.reset()
            local_bus = EventBus()
            # 订阅
            local_bus.subscribe(QUOTATION_COMPLETED, warehouse.on_quotation_completed)
            local_bus.subscribe("stock.below_safety", purchasing.on_stock_low)

            # 启动总线（后台）
            local_bus._running = True
            bus_task = asyncio.create_task(local_bus.start())

            # 发布事件
            await local_bus.publish(event["event_type"], event["payload"])
            try:
                from ws_server import publish_ws
                publish_ws(event["event_type"], event["payload"])
            except Exception:
                pass

            # 等待处理完成（最多3秒）
            await asyncio.sleep(1.5)

            # 停止总线
            local_bus.stop()
            bus_task.cancel()
            try:
                await bus_task
            except asyncio.CancelledError:
                pass

            return {
                "warehouse_alerts": warehouse.get_last_alerts(),
                "purchasing_suggestions": purchasing.get_suggestions()
            }

        loop = asyncio.new_event_loop()
        chain_result = loop.run_until_complete(run_chain())
        loop.close()

        # 返回完整事件链结果
        return jsonify({
            "ok": True,
            "event_id": event["id"],
            "quotation_id": quote_id,
            "chain": {
                "step1_deducted": len(items),
                "step2_alerts": len(chain_result["warehouse_alerts"]),
                "step3_suggestions": len(chain_result["purchasing_suggestions"]),
                "alerts": chain_result["warehouse_alerts"],
                "suggestions": chain_result["purchasing_suggestions"]
            }
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)})


# ─── 智能体状态 ───
@app.route("/api/agents/status")
@require_role("dashboard")
def agents_status():
    """返回所有智能体的运行状态"""
    try:
        from agents.quotation_agent import QuotationAgent
        from agents.trade_agent import TradeAgent
        from agents.warehouse_agent import WarehouseAgent
        from agents.purchasing_agent import PurchasingAgent

        quotation = QuotationAgent()
        trade = TradeAgent()
        warehouse = WarehouseAgent()
        purchasing = PurchasingAgent()

        agents = [
            {
                "name": quotation.name,
                "status": quotation.status,
                "type": "quotation",
                "description": "报价智能体 - 询盘解析、配件匹配、四模式算价、报价单生成"
            },
            {
                "name": trade.name,
                "status": trade.status,
                "type": "trade",
                "description": "外贸智能体 - 询盘翻译、信息提取、PI/装箱单生成"
            },
            {
                "name": warehouse.name,
                "status": warehouse.status,
                "type": "warehouse",
                "description": "仓储智能体 - 管理库存，响应报价事件减库存"
            },
            {
                "name": purchasing.name,
                "status": purchasing.status,
                "type": "purchasing",
                "description": "采购智能体 - 监听库存预警，生成补货建议"
            }
        ]

        return jsonify({
            "ok": True,
            "agents": agents,
            "count": len(agents)
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


# ═══════════ 别名映射 API ═══════════

@app.route("/api/aliases")
@require_role("quote")
def search_aliases():
    """GET /api/aliases?q=xxx — 搜索别名（FTS5 + 别名表）"""
    q = request.args.get("q", "")
    limit = request.args.get("limit", 5, type=int)
    if not q:
        return jsonify({"ok": False, "error": "缺少搜索词 q"})
    try:
        from core.aliases import search_alias
        results = search_alias(q, limit=limit)
        return jsonify({"ok": True, "data": results, "count": len(results)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e), "data": []})


@app.route("/api/aliases/suggest", methods=["POST"])
@require_role("quote")
def suggest_aliases():
    """
    POST /api/aliases/suggest — AI 建议
    Body: {"new_name": "...", "oe_number": "..."}
    返回：候选配件列表（可能是别名映射）
    """
    try:
        data = request.get_json(force=True)
        new_name = data.get("new_name", "")
        oe_number = data.get("oe_number", "")
        if not new_name and not oe_number:
            return jsonify({"ok": False, "error": "至少提供 new_name 或 oe_number"})
        from core.aliases import suggest_alias
        candidates = suggest_alias(new_name, oe_number)
        return jsonify({"ok": True, "data": candidates, "count": len(candidates)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e), "data": []})


@app.route("/api/aliases/confirm", methods=["POST"])
@require_role("quote")
def confirm_alias_api():
    """
    POST /api/aliases/confirm — 人工确认别名
    Body: {"alias_id": N} 或 {"part_id": N, "alias_text": "...", "lang": "zh"}
    如果是新增+确认：传入 part_id + alias_text，会自动创建并确认
    """
    try:
        data = request.get_json(force=True)
        from core.aliases import add_alias, confirm_alias, get_aliases

        alias_id = data.get("alias_id")
        if alias_id:
            # 直接确认已有别名
            result = confirm_alias(int(alias_id))
            if result:
                return jsonify({"ok": True, "data": result, "message": "别名已确认"})
            else:
                return jsonify({"ok": False, "error": f"别名 id={alias_id} 不存在"})

        part_id = data.get("part_id")
        alias_text = data.get("alias_text")
        if part_id and alias_text:
            lang = data.get("lang", "zh")
            # 创建并直接确认
            alias = add_alias(int(part_id), alias_text, lang, confirmed=True)
            return jsonify({"ok": True, "data": alias, "message": "别名已创建并确认"})

        return jsonify({"ok": False, "error": "请提供 alias_id 或 (part_id + alias_text)"})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)})


@app.route("/api/aliases/<int:part_id>", methods=["GET"])
@require_role("quote")
def get_part_aliases(part_id):
    """GET /api/aliases/<part_id> — 获取某配件的所有别名"""
    try:
        confirmed_only = request.args.get("confirmed_only", "0") == "1"
        from core.aliases import get_aliases
        aliases = get_aliases(part_id, confirmed_only=confirmed_only)
        return jsonify({"ok": True, "data": aliases, "count": len(aliases)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e), "data": []})


@app.route("/api/aliases/unconfirmed", methods=["GET"])
@require_role("quote")
def get_unconfirmed():
    """GET /api/aliases/unconfirmed — 获取所有待确认的 AI 建议别名"""
    try:
        from core.aliases import get_unconfirmed_aliases
        aliases = get_unconfirmed_aliases()
        return jsonify({"ok": True, "data": aliases, "count": len(aliases)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e), "data": []})


# ─── 翻译接口 ───
@app.route("/api/translate", methods=["POST"])
@require_role("quote")
def translate_text():
    """POST /api/translate — 翻译询盘文本
    请求: {"text": "...", "source_lang": "ru", "target_lang": "zh"}
    返回: {"ok": true, "translated": "..."}
    """
    try:
        data = request.get_json(force=True)
        text = data.get("text", "")
        source_lang = data.get("source_lang", "auto")
        target_lang = data.get("target_lang", "zh")

        if not text.strip():
            return jsonify({"ok": False, "error": "文本为空"})

        from oracle.client import translate as do_translate
        result = do_translate(text, target_lang=target_lang, source_lang=source_lang)

        return jsonify({"ok": True, "translated": result, "source_lang": source_lang, "target_lang": target_lang})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)})


# ─── PI 生成接口 ───
@app.route("/api/pi/generate", methods=["POST"])
@require_role("quote")
def generate_pi():
    """POST /api/pi/generate — 根据报价结果生成 PI 草稿
    请求: {
        "inquiry_text": "原文询盘",
        "items": [{"oe_number":"xxx","name":"xxx","quantity":1,"unit_price":100,"total":500}],
        "customer_name": "客户名",
        "trade_term": "FOB",
        "payment_term": "prepaid",
        "currency": "USD"
    }
    返回: {"ok": true, "pi_draft": "..."}
    """
    try:
        data = request.get_json(force=True)
        inquiry_text = data.get("inquiry_text", "")
        items = data.get("items", [])
        customer_name = data.get("customer_name", "")
        trade_term = data.get("trade_term", "FOB")
        payment_term = data.get("payment_term", "prepaid")
        currency = data.get("currency", "USD")

        if not items:
            return jsonify({"ok": False, "error": "items 为空"})

        from oracle.client import generate_pi as do_generate_pi
        pi_draft = do_generate_pi(
            inquiry_text=inquiry_text,
            items=items,
            customer_name=customer_name,
            trade_term=trade_term,
            payment_term=payment_term,
            currency=currency,
        )

        return jsonify({"ok": True, "pi_draft": pi_draft})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)})


# ─── 询盘处理（翻译+提取+PI一体化）───
@app.route("/api/inquiry/process", methods=["POST"])
@require_role("quote")
def process_inquiry():
    """POST /api/inquiry/process — 一站式询盘处理
    请求: {"text": "俄文/英文询盘", "language": "ru", "customer": "客户名"}
    返回: {"ok": true, "data": {translated_text, extracted_items, pi_draft}}
    """
    try:
        data = request.get_json(force=True)
        text = data.get("text", "")
        lang = data.get("language", "auto")
        customer = data.get("customer", "")

        if not text.strip():
            return jsonify({"ok": False, "error": "询盘文本为空"})

        import asyncio
        from agents.trade_agent import trade_agent

        # Build event
        event = {
            "id": f"API-INQUIRY-{id(text)}",
            "event_type": "inquiry.received",
            "payload": {
                "text": text,
                "language": lang,
                "customer": customer,
            }
        }

        # Run synchronously
        loop = asyncio.new_event_loop()
        result = loop.run_until_complete(trade_agent.on_inquiry(event))
        loop.close()

        return jsonify({"ok": True, "data": result})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)})




# ─── 价格变更接口 ───
@app.route("/api/price/update", methods=["POST"])
@require_role("manage")
def update_price():
    """POST /api/price/update — 更新配件价格并发布 price.changed 事件
    请求: {"oe_number": "PART-0001", "new_price": 4500.00, "brand": "A2080"}
    返回: {"ok": true, "data": {...}, "event_id": "..."}
    """
    try:
        data = request.get_json(force=True)
        oe_number = data.get("oe_number", "").strip()
        new_price = data.get("new_price")
        brand = data.get("brand", "").strip()

        if not oe_number:
            return jsonify({"ok": False, "error": "缺少 oe_number"})
        if new_price is None:
            return jsonify({"ok": False, "error": "缺少 new_price"})
        try:
            new_price = float(new_price)
        except (TypeError, ValueError):
            return jsonify({"ok": False, "error": "new_price 必须是数字"})

        from core import get_db, get_part_by_oe
        import asyncio
        import time
        import uuid

        # 1) 更新 parts 表
        db = get_db()
        # 先查旧值
        old = db.execute(
            "SELECT oe_number, list_price, brand_channel, cost_with_tax FROM parts WHERE oe_number = ?",
            (oe_number,)
        ).fetchone()

        if not old:
            db.close()
            return jsonify({"ok": False, "error": f"配件 {oe_number} 不存在"})

        old_price = old["list_price"]
        old_brand = old["brand_channel"]
        cost_with_tax = old["cost_with_tax"] or 0.0

        # 更新价格
        if brand:
            db.execute(
                "UPDATE parts SET list_price = ?, brand_channel = ? WHERE oe_number = ?",
                (new_price, brand, oe_number)
            )
        else:
            db.execute(
                "UPDATE parts SET list_price = ? WHERE oe_number = ?",
                (new_price, oe_number)
            )
        db.commit()
        db.close()

        # 2) 发布 price.changed 事件
        event = {
            "id": uuid.uuid4().hex[:8],
            "event_type": "price.changed",
            "payload": {
                "oe_number": oe_number,
                "old_price": old_price,
                "new_price": new_price,
                "brand": brand or old_brand,
                "cost_with_tax": cost_with_tax,
            },
            "timestamp": time.time()
        }

        # 在临时 event loop 中执行
        from bus.eventbus import EventBus, PRICE_CHANGED

        async def run_price_change():
            EventBus.reset()
            local_bus = EventBus()

            # 导入并注册 price.changed handler
            from bus.handlers import on_price_changed
            local_bus.subscribe(PRICE_CHANGED, on_price_changed)

            local_bus._running = True
            bus_task = asyncio.create_task(local_bus.start())

            await local_bus.publish(event["event_type"], event["payload"])
            try:
                from ws_server import publish_ws
                publish_ws(event["event_type"], event["payload"])
            except Exception:
                pass
            await asyncio.sleep(0.5)  # 等待异步处理

            local_bus.stop()
            bus_task.cancel()
            try:
                await bus_task
            except asyncio.CancelledError:
                pass

        loop = asyncio.new_event_loop()
        loop.run_until_complete(run_price_change())
        loop.close()

        return jsonify({
            "ok": True,
            "data": {
                "oe_number": oe_number,
                "old_price": old_price,
                "new_price": new_price,
                "brand": brand or old_brand,
            },
            "event_id": event["id"],
            "message": "价格已更新，price.changed 事件已发布"
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)})

# ─── 首页看板 ───
@app.route("/api/dashboard")
@require_role("dashboard")
def dashboard():
    """返回首页看板数据：核心指标 + 库存预警 + 最近活动"""
    try:
        from core import get_db
        from ledger.trace import audit
        conn = get_db()

        # 核心指标
        parts_count = conn.execute("SELECT COUNT(*) FROM parts").fetchone()[0]
        stock_count = conn.execute("SELECT COUNT(*) FROM stock").fetchone()[0]
        alert_count = conn.execute(
            "SELECT COUNT(*) FROM stock WHERE quantity <= safety_line"
        ).fetchone()[0]
        brands = conn.execute(
            "SELECT COUNT(DISTINCT brand_channel) FROM parts WHERE brand_channel != ''"
        ).fetchone()[0]
        pricing_modes = conn.execute(
            "SELECT pricing_mode, COUNT(*) as cnt FROM parts "
            "GROUP BY pricing_mode ORDER BY cnt DESC"
        ).fetchall()

        # 最近报价
        recent_quotes = conn.execute("""
            SELECT q.id, q.created_at, q.total_amount, q.status, q.customer_name,
                   COUNT(ql.id) as line_count
            FROM quotations q
            LEFT JOIN quotation_lines ql ON ql.quotation_id = q.id
            GROUP BY q.id
            ORDER BY q.created_at DESC LIMIT 5
        """).fetchall()

        # 库存预警列表
        alerts = conn.execute("""
            SELECT s.part_oe, s.quantity, s.safety_line, p.name_cn, p.brand_channel, p.list_price
            FROM stock s LEFT JOIN parts p ON s.part_oe = p.oe_number
            WHERE s.quantity <= s.safety_line
            ORDER BY (s.safety_line - s.quantity) DESC
            LIMIT 10
        """).fetchall()

        conn.close()

        # 最近活动（从审计日志）
        trace_ids = audit.get_all_quotes()
        recent_activity = trace_ids[:10] if trace_ids else []

        return jsonify({"ok": True, "data": {
            "stats": {
                "parts": parts_count,
                "stock": stock_count,
                "alerts": alert_count,
                "brands": brands,
                "modes": [{"mode": m[0], "count": m[1]} for m in pricing_modes],
            },
            "recent_quotes": [{
                "id": q["id"], "created_at": q["created_at"] or "",
                "total_amount": q["total_amount"] or 0,
                "item_count": q["line_count"] or 0,
                "status": q["status"] or "draft",
                "customer_name": q["customer_name"] or "",
            } for q in recent_quotes],
            "stock_alerts": [{
                "part_oe": a["part_oe"], "name_cn": a["name_cn"] or "",
                "quantity": a["quantity"], "safety_line": a["safety_line"],
                "brand": a["brand_channel"] or "",
                "list_price": a["list_price"] or 0,
                "gap": a["safety_line"] - a["quantity"],
            } for a in alerts],
            "recent_activity": recent_activity,
        }})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)})

# ─── 套餐查询 ───
@app.route("/api/plans")
@require_auth
def get_plans():
    """返回四个订阅套餐"""
    from config import get_plans as _load_plans
    plans = _load_plans()
    return jsonify({"ok": True, "data": plans})

# ─── 订单创建 ───
@app.route("/api/orders", methods=["GET","POST"])
@require_role("manage")
def orders():
    if request.method == "GET":
        try:
            from core import get_db
            conn = get_db()
            rows = conn.execute("SELECT * FROM orders ORDER BY created_at DESC LIMIT 20").fetchall()
            conn.close()
            data = [dict(r) for r in rows]
            return jsonify({"ok":True,"data":data,"count":len(data)})
        except Exception as e:
            return jsonify({"ok":True,"data":[],"count":0})
    
    # POST - create order
    try:
        data = request.get_json(force=True)
        plan_id = data.get("plan_id","starter")
        period = data.get("period","month")  # month or year
        
        # Get plan
        plans_list = _load_plans()
        plan = next((p for p in plans_list if p.get("id")==plan_id), {})
        amount = plan["yearly_price"] if period=="year" else plan["monthly_price"]
        
        import uuid, time
        order_id = f"ORD-{int(time.time())}-{uuid.uuid4().hex[:4]}"
        
        from core import get_db
        conn = get_db()
        conn.execute("""CREATE TABLE IF NOT EXISTS orders (
            id TEXT PRIMARY KEY, plan_id TEXT, period TEXT,
            amount REAL, status TEXT DEFAULT 'pending',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            paid_at TIMESTAMP
        )""")
        conn.execute("INSERT INTO orders (id,plan_id,period,amount,status) VALUES (?,?,?,?,'pending')",
                     (order_id, plan_id, period, amount))
        conn.commit()
        conn.close()
        
        return jsonify({"ok":True,"data":{"order_id":order_id,"amount":amount,"status":"pending"}})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"ok":False,"error":str(e)})

# ─── 支付回调(模拟) ───
@app.route("/api/orders/pay", methods=["POST"])
@require_auth
def pay_order():
    try:
        data = request.get_json(force=True)
        order_id = data.get("order_id","")
        
        from core import get_db
        conn = get_db()
        order = conn.execute("SELECT * FROM orders WHERE id=?",(order_id,)).fetchone()
        if not order:
            conn.close()
            return jsonify({"ok":False,"error":"订单不存在"})
        
        conn.execute("UPDATE orders SET status='paid', paid_at=CURRENT_TIMESTAMP WHERE id=?",(order_id,))
        
        # Update plan in a config table
        conn.execute("""CREATE TABLE IF NOT EXISTS subscription (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            plan_id TEXT, period TEXT, quota_total INTEGER,
            quota_used INTEGER DEFAULT 0,
            started_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            expires_at TIMESTAMP
        )""")
        
        plan_id = order["plan_id"]
        period = order["period"]
        plans_list = [
            {"id":"starter","quota":300},
            {"id":"standard","quota":1000},
            {"id":"unlimited","quota":3000},
            {"id":"enterprise","quota":999999},
        ]
        plan = next((p for p in plans_list if p["id"]==plan_id), plans_list[0])
        
        import datetime
        now = datetime.datetime.now()
        if period=="year":
            expires = now + datetime.timedelta(days=365)
        else:
            expires = now + datetime.timedelta(days=30)
        
        conn.execute("DELETE FROM subscription")  # Simple: one active sub
        conn.execute("INSERT INTO subscription (plan_id,period,quota_total,expires_at) VALUES (?,?,?,?)",
                     (plan_id, period, plan["quota"], expires.isoformat()))
        conn.commit()
        conn.close()
        
        return jsonify({"ok":True,"message":"支付成功","plan":plan_id,"expires":expires.isoformat()[:10]})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"ok":False,"error":str(e)})

# ─── 用量查询 ───
@app.route("/api/usage")
@require_role("dashboard")
def get_usage():
    try:
        from core import get_db
        conn = get_db()
        
        # Current subscription
        sub = conn.execute("SELECT * FROM subscription ORDER BY id DESC LIMIT 1").fetchone()
        conn.close()
        
        if not sub:
            return jsonify({"ok":True,"data":{"plan":"none","quota_total":0,"quota_used":0,"remaining":0,"expires":""}})
        
        data = {
            "plan": sub["plan_id"],
            "period": sub["period"],
            "quota_total": sub["quota_total"],
            "quota_used": sub["quota_used"] or 0,
            "remaining": max(0, (sub["quota_total"] or 0) - (sub["quota_used"] or 0)),
            "started": sub["started_at"] or "",
            "expires": sub["expires_at"] or "",
        }
        return jsonify({"ok":True,"data":data})
    except Exception as e:
        return jsonify({"ok":True,"data":{"plan":"free","quota_total":50,"quota_used":0,"remaining":50,"expires":""}})

# ─── 调用计数(报价时自动+1) ───
@app.route("/api/usage/increment", methods=["POST"])
@require_role("dashboard")
def increment_usage():
    try:
        data = request.get_json(force=True)
        count = data.get("count", 1)
        
        from core import get_db
        conn = get_db()
        conn.execute("CREATE TABLE IF NOT EXISTS usage_log (id INTEGER PRIMARY KEY AUTOINCREMENT, quotation_id TEXT, call_count INTEGER, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)")
        conn.execute("INSERT INTO usage_log (quotation_id,call_count) VALUES (?,?)", (data.get("quotation_id",""), count))
        
        conn.execute("CREATE TABLE IF NOT EXISTS subscription (id INTEGER PRIMARY KEY AUTOINCREMENT, plan_id TEXT, period TEXT, quota_total INTEGER, quota_used INTEGER DEFAULT 0, started_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP, expires_at TIMESTAMP)")
        conn.execute("UPDATE subscription SET quota_used = COALESCE(quota_used,0) + ? WHERE id = (SELECT MAX(id) FROM subscription)", (count,))
        conn.commit()
        conn.close()
        
        return jsonify({"ok":True,"incremented":count})
    except Exception as e:
        return jsonify({"ok":False,"error":str(e)})

# ═══════════ 管理后台 ═══════════
@app.route("/api/admin/stats")
@require_role("manage")
def admin_stats():
    """营收总览"""
    try:
        from core import get_db
        conn = get_db()
        
        # 订单统计
        orders = conn.execute("SELECT COUNT(*) as total, SUM(CASE WHEN status='paid' THEN amount ELSE 0 END) as revenue, COUNT(CASE WHEN status='paid' THEN 1 END) as paid FROM orders").fetchone()
        
        # 订阅统计
        subs = conn.execute("SELECT plan_id, COUNT(*) as cnt FROM orders WHERE status='paid' GROUP BY plan_id").fetchall()
        
        # 用量统计
        usage = conn.execute("SELECT COALESCE(SUM(call_count),0) as total_calls FROM usage_log").fetchone()
        
        # 系统数据
        parts = conn.execute("SELECT COUNT(*) FROM parts").fetchone()[0]
        stock = conn.execute("SELECT COUNT(*) FROM stock").fetchone()[0]
        alerts = conn.execute("SELECT COUNT(*) FROM stock WHERE quantity <= safety_line").fetchone()[0]
        quotations = conn.execute("SELECT COUNT(*) FROM quotations").fetchone()[0]
        audit_count = conn.execute("SELECT COUNT(*) FROM audit_log").fetchone()[0]
        
        conn.close()
        
        return jsonify({"ok": True, "data": {
            "revenue": {
                "total_orders": orders["total"] or 0,
                "paid_orders": orders["paid"] or 0,
                "total_revenue": orders["revenue"] or 0,
            },
            "plans": [{"plan": s["plan_id"], "count": s["cnt"]} for s in subs],
            "usage": {"total_calls": usage["total_calls"] or 0},
            "system": {
                "parts": parts, "stock": stock, "alerts": alerts,
                "quotations": quotations, "audit_logs": audit_count,
            }
        }})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e), "data": {
            "revenue": {"total_orders":0,"paid_orders":0,"total_revenue":0},
            "plans": [], "usage": {"total_calls":0},
            "system": {"parts":168,"stock":52,"alerts":7,"quotations":0,"audit_logs":1}
        }})

@app.route("/api/admin/health")
@require_role("manage")
def admin_health():
    """系统健康检查 — 无需psutil"""
    import os
    
    cpu_pct = 0
    try:
        if hasattr(os, 'getloadavg'):
            load = os.getloadavg()[0]
            ncpu = os.cpu_count() or 1
            cpu_pct = round(min(load / ncpu * 100, 100), 1)
    except:
        pass
    
    mem_pct = 0
    try:
        total = avail = 0
        for line in open('/proc/meminfo'):
            if line.startswith('MemTotal:'): total = int(line.split()[1])
            if line.startswith('MemAvailable:'): avail = int(line.split()[1])
        if total > 0: mem_pct = round((1 - avail / total) * 100, 1)
    except:
        pass
    
    disk_pct = 0
    try:
        s = os.statvfs('/')
        if s.f_blocks > 0: disk_pct = round((1 - s.f_bavail / s.f_blocks) * 100, 1)
    except:
        pass
    
    services = {}
    for svc in ["atlas-api","atlas-portal","enong-bridge"]:
        r = os.popen("systemctl is-active " + svc + " 2>/dev/null").read().strip()
        services[svc] = r == "active"
    
    uptime_str = "N/A"
    try:
        secs = int(float(open('/proc/uptime').read().split()[0]))
        d, h = divmod(secs, 86400)
        h, m = divmod(h, 3600)
        parts = []
        if d: parts.append(str(d) + "天")
        if h: parts.append(str(h) + "小时")
        parts.append(str(m // 60) + "分钟")
        uptime_str = " ".join(parts)
    except:
        pass
    
    return jsonify({"ok": True, "data": {
        "server": {"cpu_percent": cpu_pct, "memory_percent": mem_pct, "disk_percent": disk_pct},
        "services": services,
        "uptime": uptime_str,
    }})



# ═══ Procurement + Users APIs ═══
@app.route("/api/procurement/orders", methods=["GET", "POST"])
@require_role("procure")
def procurement_orders():
    import sqlite3
    if request.method == "GET":
        try:
            conn = sqlite3.connect(str(DB_PATH))
            conn.row_factory = sqlite3.Row
            rows = conn.execute("SELECT * FROM procurement_orders WHERE company_id=? ORDER BY created_at DESC LIMIT 50", (g.company_id,)).fetchall()
            data = []
            for r in rows:
                d = dict(r)
                lc = conn.execute("SELECT COUNT(*) FROM procurement_lines WHERE order_id=?", (d["id"],)).fetchone()
                d["line_count"] = lc[0] if lc else 0
                d["units"] = calc_procurement_units(d.get("sku_count", 0))
                data.append(d)
            conn.close()
            return jsonify({"ok": True, "data": data, "count": len(data)})
        except Exception as e:
            return jsonify({"ok": False, "error": str(e)})
    try:
        data = request.get_json(force=True)
        items = data.get("items", [])
        supplier = data.get("supplier", "")
        if not items:
            return jsonify({"ok": False, "error": "items required"})
        import uuid, time
        order_id = f"PO-{int(time.time())}-{uuid.uuid4().hex[:6]}"
        sku_count = len(items)
        conn = sqlite3.connect(str(DB_PATH))
        total = sum(it.get("quantity",1)*it.get("unit_price",0) for it in items)
        conn.execute("INSERT INTO procurement_orders (id,company_id,created_by,supplier,sku_count,total_amount,approver,status) VALUES (?,?,?,?,?,?,?,?)",
                     (order_id, g.company_id, g.user.get("id",""), supplier, sku_count, round(total,2), g.user.get("name",""), "draft"))
        for it in items:
            conn.execute("INSERT INTO procurement_lines (order_id,oe_number,name_cn,quantity,unit_price,supplier_name) VALUES (?,?,?,?,?,?)",
                         (order_id, it.get("oe_number",""), it.get("name_cn",""), it.get("quantity",1), it.get("unit_price",0), it.get("supplier_name",supplier)))
        conn.commit()
        units = calc_procurement_units(sku_count)
        log_usage(g.company_id, g.user.get("id",""), "procure", units)
        conn.close()
        return jsonify({"ok": True, "data": {"order_id": order_id, "sku_count": sku_count, "units": units, "total": round(total,2)}})
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"ok": False, "error": str(e)})

@app.route("/api/procurement/orders/<order_id>", methods=["GET", "PATCH"])
@require_role("procure")
def procurement_order_detail(order_id):
    import sqlite3
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    if request.method == "GET":
        order = conn.execute("SELECT * FROM procurement_orders WHERE id=?", (order_id,)).fetchone()
        if not order:
            conn.close()
            return jsonify({"ok": False, "error": "not found"})
        lines = conn.execute("SELECT * FROM procurement_lines WHERE order_id=?", (order_id,)).fetchall()
        conn.close()
        od = dict(order)
        od["units"] = calc_procurement_units(od.get("sku_count",0))
        return jsonify({"ok": True, "data": {"order": od, "lines": [dict(l) for l in lines]}})
    try:
        data = request.get_json(force=True)
        st = data.get("status","")
        if st in ("draft","pending_approval","approved","ordered","received","cancelled"):
            conn.execute("UPDATE procurement_orders SET status=? WHERE id=?", (st, order_id))
            conn.commit()
        conn.close()
        return jsonify({"ok": True, "status": st})
    except Exception as e:
        conn.close()
        return jsonify({"ok": False, "error": str(e)})

@app.route("/api/users", methods=["GET", "POST"])
@require_role("manage")
def manage_users():
    if request.method == "GET":
        users = list_users(g.company_id)
        return jsonify({"ok": True, "data": users, "count": len(users)})
    try:
        data = request.get_json(force=True)
        email = (data.get("email") or "").strip()
        if not email: return jsonify({"ok": False, "error": "email required"})
        role = data.get("role", "quoter")
        u = create_user(email, data.get("password","123456"), data.get("name",""), role, g.company_id)
        return jsonify({"ok": True, "data": u})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route("/api/users/<user_id>", methods=["DELETE"])
@require_role("manage")
def delete_user_api(user_id):
    try:
        delete_user(user_id)
        return jsonify({"ok": True, "message": "deleted"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})

@app.route("/api/roles")
def list_roles():
    return jsonify({"ok": True, "data": [{"id": k, "label": v["label"], "level": v["level"], "permissions": v["can"]} for k, v in ROLES.items()]})

@app.route("/api/me")
@require_auth
def get_me():
    u = g.user
    return jsonify({"ok": True, "data": {"id": u.get("id",""), "name": u.get("name",""), "email": u.get("email",""), "role": g.role, "role_label": get_role_label(g.role), "company_id": g.company_id}})


# ═══ SUPER ADMIN — Multi-tenant ═══
@app.route("/api/super/stats")
@require_role("platform")
def super_stats():
    import sqlite3
    try:
        conn = sqlite3.connect(str(DB_PATH))
        c = conn.cursor()
        companies = c.execute("SELECT COUNT(*) FROM companies").fetchone()[0]
        users = c.execute("SELECT COUNT(*) FROM users WHERE status='active'").fetchone()[0]
        rev = c.execute("SELECT COALESCE(SUM(amount),0) FROM orders WHERE status='paid'").fetchone()[0]
        parts = 0
        try: parts = c.execute("SELECT COUNT(*) FROM parts").fetchone()[0]
        except: pass
        quotes = c.execute("SELECT COUNT(*) FROM quotations").fetchone()[0]
        proc = c.execute("SELECT COUNT(*) FROM procurement_orders").fetchone()[0]
        conn.close()
        return jsonify({"ok":True,"data":{"companies":companies,"users":users,"revenue":rev,"parts":parts,"quotations":quotes,"procurement_orders":proc}})
    except Exception as e:
        return jsonify({"ok":False,"error":str(e)})

@app.route("/api/super/companies")
@require_role("platform")
def super_companies():
    import sqlite3
    try:
        conn = sqlite3.connect(str(DB_PATH))
        rows = conn.execute("SELECT c.*, COUNT(DISTINCT u.id) as user_count FROM companies c LEFT JOIN users u ON u.company_id=c.id GROUP BY c.id ORDER BY c.created_at DESC").fetchall()
        conn.close()
        return jsonify({"ok":True,"data":[{"id":r[0],"name":r[1],"plan_id":r[2],"subscription_end":r[3],"invite_code":r[4],"created_at":r[5],"user_count":r[6]} for r in rows],"count":len(rows)})
    except Exception as e:
        return jsonify({"ok":False,"error":str(e)})

@app.route("/api/super/companies", methods=["POST"])
@require_role("platform")
def super_create_company():
    import sqlite3, uuid, time, subprocess
    try:
        data = request.get_json(force=True)
        name = data.get("name","").strip()
        slug = data.get("slug","").strip()
        plan = data.get("plan","starter")
        admin_email = data.get("admin_email","").strip()
        admin_pass = data.get("admin_password","admin123")
        if not name or not slug:
            return jsonify({"ok":False,"error":"name and slug required"})
        cid = f"COMPANY-{slug.upper()}-{uuid.uuid4().hex[:6]}"
        conn = sqlite3.connect(str(DB_PATH))
        conn.execute("INSERT INTO companies (id,name,plan_id) VALUES (?,?,?)",(cid,name,plan))
        conn.commit(); conn.close()
        uid = create_user(admin_email, admin_pass, name+"管理员", "boss", cid)
        url = ""
        try:
            subprocess.run(["sudo","python3","/srv/atlas/scripts/deploy_client.py",slug,name,admin_email],capture_output=True,timeout=30)
            url = f"https://{slug}.traceclaw.cn"
        except: url = f"https://enter.traceclaw.cn"
        return jsonify({"ok":True,"data":{"company_id":cid,"name":name,"slug":slug,"admin_user_id":uid.get("id",""),"url":url}})
    except Exception as e:
        return jsonify({"ok":False,"error":str(e)})

@app.route("/api/super/company/<cid>")
@require_role("platform")
def super_company_detail(cid):
    import sqlite3
    try:
        conn = sqlite3.connect(str(DB_PATH))
        conn.row_factory = sqlite3.Row
        company = conn.execute("SELECT * FROM companies WHERE id=?",(cid,)).fetchone()
        if not company: conn.close(); return jsonify({"ok":False,"error":"not found"})
        users = conn.execute("SELECT id,name,email,role,status FROM users WHERE company_id=? AND status='active'",(cid,)).fetchall()
        quotes = conn.execute("SELECT id,total_amount,status,created_at FROM quotations ORDER BY created_at DESC LIMIT 10").fetchall()
        po = conn.execute("SELECT id,supplier,sku_count,total_amount,status FROM procurement_orders WHERE company_id=? LIMIT 10",(cid,)).fetchall()
        conn.close()
        return jsonify({"ok":True,"data":{"company":dict(company),"users":[dict(u) for u in users],"recent_quotes":[dict(q) for q in quotes],"procurement_orders":[dict(p) for p in po]}})
    except Exception as e:
        return jsonify({"ok":False,"error":str(e)})

@app.route("/api/super/company/<cid>", methods=["PATCH"])
@require_role("platform")
def super_update_company(cid):
    import sqlite3
    try:
        data = request.get_json(force=True)
        plan = data.get("plan_id","")
        sub_end = data.get("subscription_end","")
        conn = sqlite3.connect(str(DB_PATH))
        if plan: conn.execute("UPDATE companies SET plan_id=? WHERE id=?",(plan,cid))
        if sub_end: conn.execute("UPDATE companies SET subscription_end=? WHERE id=?",(sub_end,cid))
        conn.commit(); conn.close()
        return jsonify({"ok":True})
    except Exception as e:
        return jsonify({"ok":False,"error":str(e)})



# ═══ 客户拼图员 API ═══

@app.route("/api/puzzler/search", methods=["POST"])
def api_puzzler_search():
    """找线索: 多渠道搜索潜在客户"""
    data = request.get_json() or {}
    industry = data.get("industry", "")
    country = data.get("country", "")
    keywords = data.get("keywords", "")
    channels = data.get("channels", None)
    if not industry or not country:
        return jsonify({"ok": False, "error": "请填写行业和目标国家"}), 400
    try:
        leads = find_leads(industry, country, keywords, channels)
        return jsonify({"ok": True, "leads": leads, "count": len(leads)})
    except Exception as e:
        return jsonify({"ok": False, "error": f"搜索失败: {str(e)}"}), 500

@app.route("/api/puzzler/bgcheck", methods=["POST"])
def api_puzzler_bgcheck():
    """拼背调: 多维度企业背景调查"""
    data = request.get_json() or {}
    company = data.get("company", "")
    website = data.get("website", "")
    country = data.get("country", "")
    if not company:
        return jsonify({"ok": False, "error": "请填写公司名称"}), 400
    try:
        result = background_check(company, website, country)
        return jsonify({"ok": True, "bgcheck": result})
    except Exception as e:
        return jsonify({"ok": False, "error": f"背调失败: {str(e)}"}), 500

@app.route("/api/puzzler/email", methods=["POST"])
def api_puzzler_email():
    """写开发信: 基于背调生成个性化邮件"""
    data = request.get_json() or {}
    company_info = data.get("company", {})
    bg_check = data.get("bgcheck", {})
    language = data.get("language", "en")
    style = data.get("style", "professional")
    if not company_info or not bg_check:
        return jsonify({"ok": False, "error": "请提供公司信息和背调数据"}), 400
    try:
        result = generate_email(company_info, bg_check, language, style)
        return jsonify({"ok": True, "email": result})
    except Exception as e:
        return jsonify({"ok": False, "error": f"生成失败: {str(e)}"}), 500

@app.route("/api/puzzler/pipeline", methods=["POST"])
def api_puzzler_pipeline():
    """硬管道: 搜客 → 匹配 → 开发信 (三阶段强制，不可跳过)"""
    data = request.get_json() or {}
    industry = data.get("industry", "")
    country = data.get("country", "")
    keywords = data.get("keywords", "")
    language = data.get("language", "en")
    channels = data.get("channels", None)
    if not industry or not country:
        return jsonify({"ok": False, "error": "请填写行业和目标国家"}), 400
    try:
        pipeline = get_pipeline()
        result = pipeline.run(industry, country, keywords, channels=channels, language=language)
        return jsonify(result.to_dict())
    except Exception as e:
        return jsonify({"ok": False, "error": f"管道执行失败: {str(e)}"}), 500

# ═══ 客户开发员 - 匹配分析 ═══
@app.route("/api/puzzler/match", methods=["POST"])
def api_puzzler_match():
    """分析客户与产品线的匹配度"""
    data = request.get_json() or {}
    lead = data.get("lead", {})
    bgcheck = data.get("bgcheck", {})
    if not lead.get("company"):
        return jsonify({"ok": False, "error": "请提供客户信息"}), 400
    try:
        result = match_analysis(lead, bgcheck)
        return jsonify({"ok": True, "match": result})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

# ═══ 客户开发员 - SMTP 配置 ═══
@app.route("/api/puzzler/config", methods=["POST"])
def api_puzzler_config():
    """配置邮件发送 SMTP"""
    data = request.get_json() or {}
    host = data.get("host", "")
    port = data.get("port", 587)
    user = data.get("user", "")
    password = data.get("password", "")
    from_name = data.get("from_name", "")
    from_email = data.get("from_email", "")
    if not host or not user or not password:
        return jsonify({"ok": False, "error": "请填写SMTP服务器、用户名和密码"}), 400
    try:
        result = configure_smtp(host, port, user, password, from_name, from_email)
        return jsonify(result)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

# ═══ 客户开发员 - 发送开发信 ═══
@app.route("/api/puzzler/send", methods=["POST"])
def api_puzzler_send():
    """发送单封或批量开发信"""
    data = request.get_json() or {}
    emails = data.get("emails", [])
    
    # 单封模式
    if not emails and data.get("to_email"):
        to_email = data.get("to_email")
        subject = data.get("subject", "")
        body = data.get("body", "")
        body_html = data.get("body_html", "")
        to_name = data.get("to_name", "")
        result = send_email_via_smtp(to_email, subject, body, to_name, body_html)
        return jsonify(result)
    
    # 批量模式
    if emails:
        result = send_bulk_emails(emails)
        return jsonify(result)
    
    return jsonify({"ok": False, "error": "请提供收件人和邮件内容"}), 400

# ═══ 客户开发员 - SMTP 状态检查 ═══
@app.route("/api/puzzler/status", methods=["GET"])
def api_puzzler_status():
    """检查 SMTP 配置状态"""
    configured = bool(SMTP_CONFIG.get("host"))
    return jsonify({
        "ok": True,
        "configured": configured,
        "host": SMTP_CONFIG.get("host", ""),
        "user": SMTP_CONFIG.get("user", ""),
        "from_name": SMTP_CONFIG.get("from_name", "")
    })

# ═══ 邮件垃圾检测 ═══
@app.route("/api/puzzler/spam-check", methods=["POST"])
def api_spam_check():
    """检测邮件内容垃圾风险"""
    data = request.get_json() or {}
    subject = data.get("subject", "")
    body = data.get("body", "")
    result = check_spam_score(subject, body)
    return jsonify({"ok": True, **result})

# ═══ 报价审查标准 API（受保护）═══
@app.route("/api/review/validate", methods=["GET"])
def api_review_validate():
    """校验审查标准完整性"""
    try:
        c = ReviewCriteria.load()
        return jsonify({
            "ok": True,
            "version": c.version,
            "brands": list(c.brands.keys()),
            "customer_tiers": list(c.customer_tiers.keys()),
            "price_floors_count": len(c.price_floors),
            "approval_flow_count": len(c.approval_flow),
            "status": "完整性校验通过"
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/review/rules", methods=["GET"])
def api_review_rules():
    """获取完整审查标准（只读）"""
    try:
        c = ReviewCriteria.load()
        return jsonify({"ok": True, "data": c.all_data})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/review/add-brand", methods=["POST"])
def api_review_add_brand():
    """新增品牌规则（只能新增，不能修改已有）"""
    data = request.get_json() or {}
    brand_code = data.get("brand_code", "")
    tiers = data.get("tiers", [])
    cap_discount = data.get("cap_discount", 0)
    special_policy = data.get("special_policy", "")
    if not brand_code or not tiers:
        return jsonify({"ok": False, "error": "请提供 brand_code 和 tiers"}), 400
    try:
        c = ReviewCriteria.load()
        c.add_brand_rule(brand_code, tiers, cap_discount, special_policy)
        c.save()
        return jsonify({"ok": True, "message": f"品牌 {brand_code} 已新增"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400

@app.route("/api/review/add-tier", methods=["POST"])
def api_review_add_tier():
    """新增客户等级"""
    data = request.get_json() or {}
    tier = data.get("tier", 0)
    rule = data.get("rule", "")
    if not tier or not rule:
        return jsonify({"ok": False, "error": "请提供 tier 和 rule"}), 400
    try:
        c = ReviewCriteria.load()
        c.add_customer_tier(int(tier), rule)
        c.save()
        return jsonify({"ok": True, "message": f"等级 {tier} 已新增"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400

# ═══ 代理配置 ═══
_proxy_config = {"enabled": False, "protocol": "socks5", "host": "", "port": 1080, "user": "", "password": ""}
_proxy_config_path = os.path.join(os.path.dirname(__file__), "proxy.json")

def _load_proxy_from_disk():
    global _proxy_config
    try:
        with open(_proxy_config_path, "r") as f:
            _proxy_config.update(json.load(f))
    except:
        pass

_load_proxy_from_disk()

@app.route("/api/config/proxy", methods=["GET"])
def api_get_proxy():
    return jsonify({"ok": True, "config": _proxy_config})

@app.route("/api/config/proxy", methods=["POST"])
def api_set_proxy():
    global _proxy_config
    data = request.get_json() or {}
    _proxy_config["enabled"] = data.get("enabled", False)
    _proxy_config["protocol"] = data.get("protocol", "socks5")
    _proxy_config["host"] = data.get("host", "")
    _proxy_config["port"] = int(data.get("port", 1080))
    _proxy_config["user"] = data.get("user", "")
    _proxy_config["password"] = data.get("password", "")
    try:
        with open(_proxy_config_path, "w") as f:
            json.dump(_proxy_config, f, indent=2)
        return jsonify({"ok": True, "message": "代理配置已保存"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/api/config/proxy/test", methods=["POST"])
def api_test_proxy():
    """测试代理是否可达"""
    import time
    if not _proxy_config.get("enabled") or not _proxy_config.get("host"):
        return jsonify({"ok": False, "error": "代理未启用或未配置"})
    try:
        proxy_url = f"{_proxy_config['protocol']}://{_proxy_config['host']}:{_proxy_config['port']}"
        proxies = {"http": proxy_url, "https": proxy_url}
        t0 = time.time()
        resp = requests.get("https://www.google.com", proxies=proxies, timeout=10)
        latency = round((time.time() - t0) * 1000)
        return jsonify({"ok": True, "latency_ms": latency, "status": resp.status_code})
    except Exception as e:
        return jsonify({"ok": False, "error": f"代理不可达: {str(e)[:100]}"})

# ═══ SMTP 测试 ═══
@app.route("/api/puzzler/smtp-test", methods=["POST"])
def api_smtp_test():
    """发送测试邮件验证 SMTP 配置"""
    try:
        from puzzler_engine import send_email_via_smtp, SMTP_CONFIG
        if not SMTP_CONFIG.get("host"):
            return jsonify({"ok": False, "error": "SMTP 未配置"})
        result = send_email_via_smtp(
            to_email=SMTP_CONFIG.get("user", ""),
            subject="[Atlas 测试] 邮件配置验证",
            body="这是一封测试邮件。\n\n如果您收到此邮件，说明 SMTP 配置正确。\n\nAtlas 客户开发引擎",
            to_name="Atlas User",
            body_html=""
        )
        return jsonify(result)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)[:200]})


if __name__ == "__main__":
    import os
    os.environ.setdefault("DEEPSEEK_API_KEY", "sk-f4aef21293b5472d9f22f86ad289b573")
    port = int(os.environ.get("ATLAS_API_PORT", 3095))

    # ─── 启动 WebSocket 事件推送服务 :3096 ───
    try:
        from ws_server import start_ws_in_thread
        ws_port = int(os.environ.get("ATLAS_WS_PORT", 3096))
        start_ws_in_thread(port=ws_port)
        print(f"🔌 WebSocket 事件推送已启动 ws://0.0.0.0:{ws_port}")
    except Exception as e:
        print(f"[warn] WebSocket 启动失败（非致命）: {e}")

    print(f"🚀 元策·擎天 API 启动 :{port}")
    app.run(host="0.0.0.0", port=port, debug=False)
