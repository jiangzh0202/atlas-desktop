"""
擎天 Atlas — 恩同报价工作簿完整解析器
吃掉四个 Sheet 的全部逻辑：客户原表 → 报价留底 → 报价策略 → quotation

真实结构（俄罗斯9520报价留底）：
┌──────────────────────────────────────────────────────────┐
│ Sheet 1: 客户原表 (Customer Order)                       │
│   客户俄文询盘：№ | Артикул(OE) | Товары(品名) | Кол-во │
│   168行配件清单，含 мин.партия(最小包装量)               │
│   输出：inquiry_lines[{oe, name_ru, qty, min_pack}]     │
├──────────────────────────────────────────────────────────┤
│ Sheet 2: 报价留底 (Internal Quotation Worksheet)         │
│   内部定价工底：品牌/供货号/中文品名/福田牌价/折扣/折后价 │
│   输出：part_records + pricing_records                   │
├──────────────────────────────────────────────────────────┤
│ Sheet 3: 报价策略 (Pricing Strategy)                     │
│   复杂的多区块布局：                                      │
│   - 左半：喷油器/滤清器/燃油泵 的库存+折扣表              │
│   - 中半：A2080品牌折扣矩阵（单价区间×数量阶梯+占比）    │
│   - 右半：卡友配品牌价格对照+顶格政策                     │
│   输出：discount_matrix, special_policies                │
├──────────────────────────────────────────────────────────┤
│ Sheet 4: quotation (Final Customer Quotation)            │
│   最终对客报价单：英俄双语，含质量等级A+/B+、FOB条款      │
│   输出：quotation_output (对标格式)                      │
└──────────────────────────────────────────────────────────┘
"""

import openpyxl
from dataclasses import dataclass, field
from typing import Optional
from pathlib import Path

def _safe_float(v, default=0.0):
    try: return float(v)
    except: return default

def _safe_qty(v):
    try: return int(str(v or "0").replace("\n", "").replace("'", "").strip())
    except: return 0

@dataclass
class InquiryLine:
    """客户询盘行"""
    line_no: int
    oe_number: str
    name_ru: str
    quantity: int
    unit: str = "шт"
    min_pack: int = 1  # мин.партия

@dataclass  
class PartRecord:
    """配件记录（从报价留底提取）"""
    oe_number: str
    name_ru: str = ""
    name_cn: str = ""
    brand: str = ""            # 品牌/渠道
    supply_number: str = ""    # 供货号
    list_price: float = 0.0    # 牌价单价
    list_price_total: float = 0.0  # 牌价总额
    discount_pct: float = 0.0  # 折扣%
    discount_coeff: float = 1.0  # 折扣系数
    discounted_price: float = 0.0  # 折后单价
    discounted_total: float = 0.0  # 折后总额
    pricing_mode: str = "STANDARD"
    fixed_price: float = 0.0
    cost_with_tax: float = 0.0
    note: str = ""
    salesperson: str = ""
    quoter: str = ""

@dataclass
class DiscountTier:
    """折扣阶梯"""
    price_min: float; price_max: float
    qty_threshold: Optional[int]
    discount_lt: float; discount_gte: float
    item_count: int = 0        # 该区间项数
    amount_ratio: float = 0.0  # 金额占比
    cap: str = ""              # 顶格标记

@dataclass
class BrandStrategy:
    """品牌策略块"""
    brand_name: str
    tiers: list = field(default_factory=list)
    cap_discount: float = 0.0
    special_notes: list = field(default_factory=list)

@dataclass
class QuotationSheet:
    """最终报价单"""
    to_name: str = ""
    from_company: str = "Shiyan Enter Auto Parts Co., Ltd."
    attn: str = ""
    date: str = ""
    lines: list = field(default_factory=list)
    total_amount: float = 0.0
    trade_term: str = "FOB"
    notes: list = field(default_factory=list)


class EnTongWorkbook:
    """
    恩同工作簿完整解析器
    用法:
        wb = EnTongWorkbook("俄罗斯9520报价留底.xlsx")
        inquiries = wb.parse_inquiry()      # Sheet 1
        parts = wb.parse_worksheet()         # Sheet 2
        strategies = wb.parse_strategies()   # Sheet 3
        quotation = wb.parse_quotation()     # Sheet 4
    """
    
    def __init__(self, path: str):
        self.wb = openpyxl.load_workbook(path, data_only=True)
        self._validate()
    
    def _validate(self):
        required = {'客户原表', '报价留底', '报价策略', 'quotation'}
        missing = required - set(self.wb.sheetnames)
        if missing:
            raise ValueError(f"缺少Sheet: {missing}")
    
    # ═══════ Sheet 1: 客户原表 ═══════
    def parse_inquiry(self) -> list[InquiryLine]:
        """解析俄文客户询盘"""
        ws = self.wb['客户原表']
        lines = []
        
        # 数据从第7行开始
        for row in ws.iter_rows(min_row=7, values_only=True):
            oe = str(row[3] or "").strip()  # col[3]=Артикул
            if not oe or oe == 'None':
                continue
            
            name_ru = str(row[7] or "").strip()  # col[7]=Товары
            qty = int(row[20] or 0)  # col[20]=Кол-во
            if qty <= 0:
                continue
            
            lines.append(InquiryLine(
                line_no=len(lines) + 1,
                oe_number=oe,
                name_ru=name_ru[:120],
                quantity=qty,
                unit=str(row[23] or "шт").strip(),
            ))
        
        return lines
    
    # ═══════ Sheet 2: 报价留底 ═══════
    def parse_worksheet(self) -> list[PartRecord]:
        """
        解析内部报价工底
        列映射（已验证）:
          col[1]=序号, col[3]=OE, col[4]=俄文名, col[5]=数量
          col[7]=单位, col[9]=品牌, col[10]=供货号
          col[12]=中文品名, col[13]=牌价总额, col[14]=牌价
          col[15]=折扣%, col[16]=折扣系数, col[17]=折后价, col[18]=折后总额
          col[19]=备注, col[20]=业务员, col[21]=报价员
        """
        ws = self.wb['报价留底']
        records = []
        
        for row in ws.iter_rows(min_row=4, values_only=True):
            oe = str(row[3] or "").strip()
            if not oe or oe == 'None':
                continue
            
            note = str(row[19] or "")
            disc_str = str(row[15] or "0").strip()
            
            # 判断定价模式
            if "库存一口价" in note or "一口价" in note or disc_str in ("/", "无", ""):
                mode = "FIXED_STOCK"
            elif "含税进价" in note:
                mode = "COST_BASED"
            elif "一单一议" in note:
                mode = "NEGOTIATED"
            else:
                mode = "STANDARD"
            
            record = PartRecord(
                oe_number=oe,
                name_ru=str(row[4] or "")[:120],
                name_cn=str(row[12] or "")[:120],
                brand=str(row[9] or "").strip(),
                supply_number=str(row[10] or "").strip(),
                list_price=_safe_float(row[14]),
                list_price_total=_safe_float(row[13]),
                discount_pct=_safe_float(row[15]),
                discount_coeff=_safe_float(row[16], 1.0),
                discounted_price=_safe_float(row[17]),
                discounted_total=_safe_float(row[18]),
                pricing_mode=mode,
                fixed_price=_safe_float(row[17]) if mode == "FIXED_STOCK" else 0,
                cost_with_tax=_safe_float(row[17]) if mode == "COST_BASED" else 0,
                note=note,
                salesperson=str(row[20] or "").strip(),
                quoter=str(row[21] or "").strip(),
            )
            records.append(record)
        
        return records
    
    # ═══════ Sheet 3: 报价策略 ═══════
    def parse_strategies(self) -> list[BrandStrategy]:
        """
        解析报价策略（最复杂的Sheet，多区块布局）
        
        布局结构:
          行1: 区块标题行 (喷油器 / 卡友配 / 一单一议)
          行2: 列标题 (供货号|品名|库存|数量|牌价|...|折扣|...|顶格)
          行3-N: 数据行
          空行: 区块分隔
          下一区块: 滤清器 / E9300 ...
          下一区块: 燃油泵 / ...
          下一区块: 东亚 / ...
        """
        ws = self.wb['报价策略']
        strategies = []
        
        # 扫描全表，按空行分区块
        current_block = []
        current_brand = ""
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), 1):
            # 检查是否空行
            has_content = any(c is not None and str(c).strip() for c in row[:10] if c)
            
            if not has_content and current_block:
                # 空行=区块结束
                strategy = self._parse_strategy_block(current_brand, current_block)
                if strategy:
                    strategies.append(strategy)
                current_block = []
                current_brand = ""
                continue
            
            if not has_content:
                continue
            
            # 判断是否是新区块标题行
            first_cell = str(row[0] or "").strip()
            if first_cell and first_cell not in ("供货号", "滤清器供货号", ""):
                if current_block and current_brand != first_cell:
                    strategy = self._parse_strategy_block(current_brand, current_block)
                    if strategy:
                        strategies.append(strategy)
                    current_block = []
                current_brand = first_cell
            
            current_block.append([str(c or "") for c in row[:28]])
        
        # 最后一个区块
        if current_block:
            strategy = self._parse_strategy_block(current_brand, current_block)
            if strategy:
                strategies.append(strategy)
        
        return strategies
    
    def _parse_strategy_block(self, brand: str, rows: list) -> Optional[BrandStrategy]:
        """解析单个策略区块"""
        if not brand or len(rows) < 2:
            return None
        
        strategy = BrandStrategy(brand_name=brand)
        
        # 从行2提取列标题确定列映射
        header_row = rows[1] if len(rows) > 1 else rows[0]
        
        # 数据行从行3开始
        for row in rows[2:]:
            # 提取折扣规则文本（通常在col[14]位置）
            rule_text = row[14] if len(row) > 14 else ""
            cap_text = row[17] if len(row) > 17 else ""
            
            if rule_text and rule_text not in ("None", ""):
                strategy.special_notes.append(rule_text)
            if cap_text and cap_text not in ("None", ""):
                strategy.cap_discount = self._parse_cap(cap_text)
        
        return strategy
    
    def _parse_cap(self, text: str) -> float:
        """解析顶格数字"""
        try: return float(text)
        except: return 0.0
    
    # ═══════ Sheet 4: quotation ═══════
    def parse_quotation(self) -> QuotationSheet:
        """解析最终对客报价单"""
        ws = self.wb['quotation']
        sheet = QuotationSheet()
        
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            text = str(row[0] or "")
            if "To.:" in text or "To." in text:
                sheet.to_name = str(row[4] or row[0] or "").replace("To.:", "").strip()
            if "Attn.:" in text:
                sheet.attn = str(row[4] or "").strip()
            if "Date:" in text:
                sheet.date = str(row[4] or "").strip()
        
        # 配件行从 Row 8 开始（Row 7 是表头）
        for row in ws.iter_rows(min_row=8, values_only=True):
            oe = str(row[1] or "").strip()  # col[1]=Part No.
            if not oe or "Total" in oe:
                if "Total" in oe:
                    try: sheet.total_amount = float(row[7])
                    except: pass
                break
            
            sheet.lines.append({
                "no": row[0], "part_no": oe,
                "name_ru": str(row[2] or "")[:120],
                "quality": str(row[3] or "A+").strip(),
                "qty": _safe_qty(row[4]),
                "unit": str(row[5] or "PC").strip(),
                "unit_price": _safe_float(row[6]),
                "total": _safe_float(row[7]),
                "remark": str(row[8] or "").strip(),
            })
        
        # 底部备注
        for row in ws.iter_rows(min_row=170, max_row=ws.max_row, values_only=True):
            text = str(row[0] or "")
            if "Price Terms" in text or "FOB" in text:
                sheet.notes.append(text)
                if "FOB" in text: sheet.trade_term = "FOB"
                elif "CIF" in text: sheet.trade_term = "CIF"
        
        return sheet


# ═══════ 测试 ═══════
if __name__ == "__main__":
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else "tests/fixtures/dmitriy_quotation.xlsx"
    
    wb = EnTongWorkbook(path)
    
    inquiries = wb.parse_inquiry()
    print(f"Sheet1 客户原表: {len(inquiries)} 条询盘")
    for i in inquiries[:3]:
        print(f"  #{i.line_no} {i.oe_number} {i.name_ru[:50]} ×{i.quantity}")
    
    parts = wb.parse_worksheet()
    print(f"\nSheet2 报价留底: {len(parts)} 条配件")
    modes = {}
    for p in parts: modes[p.pricing_mode] = modes.get(p.pricing_mode, 0) + 1
    print(f"  定价模式: {modes}")
    
    strategies = wb.parse_strategies()
    print(f"\nSheet3 报价策略: {len(strategies)} 个策略区块")
    for s in strategies:
        print(f"  {s.brand_name}: {len(s.special_notes)}条规则, 顶格={s.cap_discount}")
    
    quote = wb.parse_quotation()
    print(f"\nSheet4 报价单: To={quote.to_name}, {len(quote.lines)}行, 总额={quote.total_amount:,.2f}")
